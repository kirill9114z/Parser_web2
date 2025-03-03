import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import re
import os

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--start-maximized")

driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)


def add_cookies(driver, cookies):
    driver.get("https://begora.ru")
    for cookie in cookies:
        driver.add_cookie(cookie)


def scroll_down(driver, scroll_pause_time=1, scroll_limit=2):
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(scroll_limit):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(scroll_pause_time)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


# 3
cookies = [
    {'name': 'PHPSESSID', 'value': 'pu2sljrn3tlm1gcojt8aj6f9k'},
    {'name': 'resellerId', 'value': '7423727'},
    {'name': 'sortOptionsSelectLinkNameSearch', 'value': 'price'},
    {'name': 'sortOptionsSelectSortDirection', 'value': '0'},
    {'name': 'visited', 'value': '1'},
    {'name': 'visited_locale', 'value': '1'},
    {"name": "cartGuestId", "value": "42208322d60afe1eb40373be694121ed"},
    {"name": "CF_VERIFIED_DEVICE_e5b4c95ab9f90c95884adf15c27e59bafb5c39011663f30743bb15008f96cec2",
     "value": "1710429905"}
]

input_file_path = 'ПРОЦЕНКА1.xlsx'
df_input = pd.read_excel(input_file_path, header=None)

marks = df_input.iloc[:, 0].tolist()
ids = df_input.iloc[:, 1].tolist()

results = []
special_items = []
best_offers = {}

previous_dir = 'C:\\Users\\эксперт\\OneDrive\\Рабочий стол\\previous_files'
new_dir = 'C:\\Users\\эксперт\\OneDrive\\Рабочий стол\\new_files'
os.makedirs(previous_dir, exist_ok=True)
os.makedirs(new_dir, exist_ok=True)

previous_file_path = os.path.join(previous_dir, 'Updated_Результаты.xlsx')
df_previous = pd.read_excel(previous_file_path, sheet_name='Special Items')

if 'row_num' not in df_previous.columns:
    df_previous['row_num'] = range(2, len(df_previous) + 2)


def extract_days(text):
    if 'часов' in text or 'часа' in text:
        hours = re.findall(r'\d+', text)
        return [int(hour) / 24 for hour in hours]
    else:
        days = re.findall(r'\d+', text)
        return [int(day) for day in days]


def has_special_title(element):
    special_titles = [
        "Рекомендуем данный склад",
        "Официальный дистрибьютор",
        "Официальный дистрибьютор или надежный поставщик",
        "Является официальным дилером"
    ]
    title = element.get_attribute('title')
    return title in special_titles


def compare_items(current_items, previous_items_df):
    comparison_results = []
    seen_items = {}

    previous_dict = previous_items_df.groupby(['mark', 'id', 'Поставщик']).apply(
        lambda df: df.drop(columns=['mark', 'id', 'Поставщик']).to_dict('records')).to_dict()

    for item in current_items:
        key = (item['mark'], item['id'], item['Поставщик'])
        if key in previous_dict:
            if key not in seen_items:
                seen_items[key] = 0
            else:
                seen_items[key] += 1

            if seen_items[key] < len(previous_dict[key]):
                previous_item = previous_dict[key][seen_items[key]]
            else:
                continue

            current_quantity = int(item['Количество']) if item['Количество'].isdigit() else 0
            previous_quantity = int(previous_item['Количество']) if str(previous_item['Количество']).isdigit() else 0

            current_price_str = re.sub(r'[^\d,\.]', '', str(item['Цена']))
            previous_price_str = re.sub(r'[^\d,\.]', '', str(previous_item['Цена']))

            current_price = float(current_price_str.replace(',', '.'))
            previous_price = float(previous_price_str.replace(',', '.'))

            quantity_change = current_quantity - previous_quantity
            price_change = current_price - previous_price

            comparison_results.append({
                'mark': item['mark'],
                'id': item['id'],
                'Поставщик': item['Поставщик'],
                'Срок поставки': item['Срок поставки'],
                'Цена': item['Цена'],
                'Количество': item['Количество'],
                'Изменение количества': f'{quantity_change:+}',
                'Изменение цены': f'{price_change:+.2f}'
            })
        else:
            comparison_results.append({
                'mark': item['mark'],
                'id': item['id'],
                'Поставщик': item['Поставщик'],
                'Срок поставки': item['Срок поставки'],
                'Цена': item['Цена'],
                'Количество': item['Количество'],
                'Изменение количества': 'отсутствует в старом файле',
                'Изменение цены': 'отсутствует в старом файле'
            })

    return comparison_results


def safe_float(value1, default=0.0):
    try:
        clean_value = re.sub(r'[^\d,\.]', '', str(value1).replace('₽', '')).replace(',', '.')
        return float(clean_value)
    except ValueError:
        return default


for mark, id in zip(marks, ids):
    url = f'https://begora.ru/search/{mark}/{id}'
    driver.get("https://begora.ru")
    add_cookies(driver, cookies)
    driver.get(url)

    scroll_down(driver, scroll_pause_time=1, scroll_limit=2)

    try:
        more_details_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'infoLink')))
        ActionChains(driver).move_to_element(more_details_button).click(more_details_button).perform()
        time.sleep(1)
    except:
        print(f"Не удалось найти кнопку 'Подробнее' для {mark}, {id}")
        continue

    properties = {}
    for i in range(2, 7):
        try:
            key = driver.find_element(By.XPATH, f'//*[@id="properties"]/table/tbody/tr[{i}]/td[1]').text.strip()
            value = driver.find_element(By.XPATH, f'//*[@id="properties"]/table/tbody/tr[{i}]/td[2]').text.strip()
            properties[key] = value
        except Exception as e:
            print(f"Ошибка при обработке свойства для {mark}, {id}: {e}")

    # Закрытие окна "Подробнее"
    try:
        close_button = driver.find_element(By.CLASS_NAME, 'fancybox-close')
        ActionChains(driver).move_to_element(close_button).click(close_button).perform()
        time.sleep(1)
    except:
        print(f"Не удалось найти кнопку закрытия для {mark}, {id}")

    found_unknown = False
    tbody_elements = driver.find_elements(By.XPATH, '//*[@id="searchResultsTable"]/tbody')
    print(f'Найдено {len(tbody_elements)} элементов <tbody> для {mark}, {id}')

    for tbody in tbody_elements:
        tr_elements = tbody.find_elements(By.XPATH, './tr')
        tr_count = len(tr_elements)

        print(f'Количество <tr> элементов внутри <tbody>: {tr_count}')

        for i in range(2, tr_count + 1):
            try:
                product_xpath = f'//*[@id="searchResultsTable"]/tbody/tr[{i}]'
                product = driver.find_element(By.XPATH, product_xpath)

                try:
                    supplier_element = product.find_element(By.XPATH, './td[6]')
                    supplier_text = supplier_element.text.strip()
                except:
                    supplier_text = "Неизвестно"
                    print(f'Не удалось найти элемент поставщика для {mark}, {id}, строка {i}')

                try:
                    delivery_time_element = product.find_element(By.XPATH, './td[10]')
                    delivery_time_text = delivery_time_element.text.strip()
                    delivery_text = extract_days(delivery_time_text)
                except:
                    delivery_time_text = "Неизвестно"
                    delivery_text = []
                    print(f'Не удалось найти элемент срока поставки для {mark}, {id}, строка {i}')

                try:
                    price_element = product.find_element(By.XPATH, './td[13]')
                    price_text = price_element.text.strip()
                    price_value = float(re.findall(r'\d+\,\d+|\d+', price_text.replace(' ', ''))[0].replace(',', '.'))
                except:
                    price_text = "Неизвестно"
                    price_value = float('inf')
                    print(f'Не удалось найти элемент цены для {mark}, {id}, строка {i}')

                try:
                    c = product.find_element(By.XPATH, './td[7]/span')
                    in_stock = c.text.strip()
                except:
                    in_stock = 'Неизвестно'
                    print(f'Не удалось найти элемент в наличие для {mark}, {id}, строка {i}')
                try:
                    description = product.find_element(By.XPATH, './td[5]')
                    description_text = description.text.strip().split()
                    if len(description_text) > 5:
                        description_text_new = ' '.join(description_text[:6])
                    else:
                        description_text_new = ' '.join(description_text)
                except:
                    print(f'Не удалось найти элемент описания для {mark}, {id}, строка {i}')

                # Проверка наличия значков в столбце "Склад" и фильтрация по title
                try:
                    warehouse_icons = product.find_elements(By.XPATH, './td[9]/*[(name()="span") or (name()="img")]')
                    if any(has_special_title(icon) for icon in warehouse_icons):
                        special_items.append({
                            'mark': mark,
                            'id': id,
                            'Поставщик': supplier_text,
                            'Срок поставки': delivery_time_text,
                            'Цена': price_text,
                            'Количество': in_stock,
                            'Склад': ", ".join([icon.get_attribute('outerHTML') for icon in warehouse_icons if
                                                has_special_title(icon)]).split('=')[-1].split('>')[0]
                        })
                except Exception as e:
                    print(f"Ошибка при проверке значков для {mark, id}, строка {i}: {e}")

                # Фильтрация товаров с максимальным сроком доставки больше 20 дней, исключая товары с доставкой в часах
                if "дней" in delivery_time_text or "дня" in delivery_time_text:
                    days = list(map(int, re.findall(r'\d+', delivery_time_text)))
                    if any(day > 20 for day in days):
                        continue

                if supplier_text == "Неизвестно" and delivery_time_text == "Неизвестно" and price_text == "Неизвестно":
                    found_unknown = True
                    break

                result = {
                    'mark': mark,
                    'id': id,
                    'Поставщик': supplier_text,
                    "Описание": description_text_new,
                    'Срок поставки': delivery_time_text,
                    'Цена': price_text,
                    'Количество': in_stock
                }

                result.update(properties)

                results.append(result)

                if id not in best_offers:
                    best_offers[id] = {
                        'cheapest': result,
                        'fastest': result
                    }
                else:
                    if price_value < float(
                            re.findall(r'\d+\,\d+|\d+', best_offers[id]['cheapest']['Цена'].replace(' ', ''))[
                                0].replace(',', '.')):
                        best_offers[id]['cheapest'] = result
                    current_fastest_days = extract_days(best_offers[id]['fastest']['Срок поставки'])
                    new_delivery_days = extract_days(delivery_time_text)
                    if min(new_delivery_days) < min(current_fastest_days):
                        best_offers[id]['fastest'] = result
            except Exception as e:
                print(f"Ошибка при обработке продукта для {mark, id}, строка {i}: {e}")
                continue
        if found_unknown:
            break
driver.quit()

df_previous['Цена'] = df_previous['Цена'].apply(safe_float)
df_previous = df_previous[df_previous['Цена'] != 0.0]

df_results = pd.DataFrame(results)
df_results['Цена'] = df_results['Цена'].apply(safe_float)
df_results = df_results[df_results['Цена'] != 0.0]

df_results_filtered = df_results.loc[df_results.groupby(['mark', 'id', 'Срок поставки', 'Количество'])['Цена'].idxmin()]
df_results_filtered.reset_index(drop=True, inplace=True)

df_special_items = pd.DataFrame(special_items)
df_special_items['Цена'] = df_special_items['Цена'].apply(safe_float)
df_special_items = df_special_items[df_special_items['Цена'] != 0.0]

df_special_items_filtered = df_special_items.loc[
    df_special_items.groupby(['mark', 'id', 'Срок поставки', 'Количество'])['Цена'].idxmin()]
df_special_items_filtered.reset_index(drop=True, inplace=True)

comparison_results = compare_items(df_special_items_filtered.to_dict('records'), df_previous)

wb = Workbook()
ws_results = wb.active
ws_results.title = 'Results'

if not df_results_filtered.empty:
    headers = list(df_results_filtered.columns)
    ws_results.append(headers)

    for _, row in df_results_filtered.iterrows():
        ws_results.append(row.tolist())

ws_best_offers = wb.create_sheet(title="Best Offers")
ws_best_offers.append(['id', 'Тип предложения'] + headers)

for id, offers in best_offers.items():
    if isinstance(offers['cheapest'], dict) and isinstance(offers['fastest'], dict):
        ws_best_offers.append([id, 'Самое дешевое'] + [offers['cheapest'].get(header, '') for header in headers])
        ws_best_offers.append([id, 'Самое быстрое'] + [offers['fastest'].get(header, '') for header in headers])
    else:
        print(f"Неправильный тип данных в best_offers для id: {id}")

ws_special_items = wb.create_sheet(title="Special Items")
if not df_special_items_filtered.empty:
    special_headers = list(df_special_items_filtered.columns)
    ws_special_items.append(special_headers)

    for _, row in df_special_items_filtered.iterrows():
        ws_special_items.append(row.tolist())

df_comparison = pd.DataFrame(comparison_results)
df_comparison = df_comparison.drop_duplicates(subset=['mark', 'id', 'Поставщик', 'Срок поставки', 'Количество'])

ws_comparison = wb.create_sheet(title="Comparison")

comparison_headers = [
    'mark', 'id', 'Поставщик', 'Срок поставки', 'Цена', 'Количество',
    'Изменение количества', 'Изменение цены'
]
ws_comparison.append(comparison_headers)

for _, row in df_comparison.iterrows():
    ws_comparison.append(row.tolist())

output_file_path = previous_file_path
wb.save(output_file_path)

print(f"Результаты сохранены в файл: {output_file_path}")
